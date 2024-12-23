#!/usr/bin/env python3

"""
Aria Corona - December 23rd, 2024
MOD_DailyReport_Excel.py
This script generates a daily report in Excel format for Meno On-Demand using data from Notion databases.
It is designed to be flexible and scalable to accommodate changes or additions to the report.
It collects data on active orders, jobs, and their statuses, and compiles it into an Excel sheet.
The report includes customer order data, product data, and job status information.
The generated report is then emailed to the specified recipients.

Modules:
    - NotionApiHelper: Helper class for interacting with the Notion API.
    - AutomatedEmails: Class for sending automated emails.
    - datetime: Module for manipulating dates and times.
    - os: Module for interacting with the operating system.
    - logging: Module for logging messages.
    - openpyxl: Module for working with Excel files.
    - json: Module for working with JSON data.

Functions:
    - query_db(db_id, filter): Queries a Notion database with the given filter and returns the response.
    - get_customer_data(data): Processes order data to generate customer-specific statistics.
    - get_product_data(data): Processes job data to generate product-specific statistics.
    - write_header(ws, orders_shipped_today, orders_shipped_this_week, job_in_error, orders_received_today): Writes the header section of the Excel report.
    - write_table(ws, data, total, headers, row_num): Writes a table of data to the Excel sheet.
    - write_tracker(ws, row_num): Writes job status information to the Excel sheet.
    - set_column_width(ws): Sets the column widths in the Excel sheet based on the maximum width of the data.
    - build_xlsx(orders_shipped_today, orders_shipped_this_week, jobs_in_error, customer_data, product_data, total_orders, total_items, orders_received_today): Builds the Excel report.
    - main(): Main function that orchestrates the data collection, processing, report generation, and email sending.

Constants:
    - OUTPUT_DIRECTORY: Directory where the output Excel file will be saved.
    - SHEET_OUTPUT_PATH: Path to the output Excel file.
    - TODAY: Current date and time.
    - TODAY_STR: Current date as a string in the format YYYY-MM-DD.
    - EMAIL_CONFIG_PATH: Path to the email configuration file.
    - SUBJECT: Subject of the automated email.
    - BODY: Body of the automated email.
    - JOB_DB_ID: Notion database ID for jobs.
    - ORDER_DB_ID: Notion database ID for orders.
    - WORKSHEET_TITLE: Title of the Excel worksheet.
    - CUSTOMER_DATA_HEADERS: Headers for the customer data table.
    - PRODUCT_DATA_HEADERS: Headers for the product data table.
    - ACTIVE_ORDERS_FILTER: Filter for querying active orders.
    - ACTIVE_JOBS_FILTER: Filter for querying active jobs.
    - ORDERS_SHIPPED_TODAY_FILTER: Filter for querying orders shipped today.
    - ORDERS_RECEIVED_TODAY_FILTER: Filter for querying orders received today.
    - ORDERS_SHIPPED_THIS_WEEK_FILTER: Filter for querying orders shipped this week.
    - JOBS_IN_ERROR_FILTER: Filter for querying jobs in error.
"""


from NotionApiHelper import NotionApiHelper
from AutomatedEmails import AutomatedEmails
from datetime import datetime, timedelta
import os, logging, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

print("Starting Daily Report...")

# Initialize Notion API helper
notion_helper = NotionApiHelper()

# Output directory
OUTPUT_DIRECTORY = "output"
SHEET_OUTPUT_PATH = os.path.join(OUTPUT_DIRECTORY, f"MOD_Daily_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)

# Logging configuration
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('logs/MOD_DailyReport_Excel.log'),
                        logging.StreamHandler()
                    ])
logger = logging.getLogger(__name__)

# Initializing timestamp
TODAY = datetime.now()
TODAY_STR = TODAY.strftime("%Y-%m-%d")

# Automated Email Configuration
automated_emails = AutomatedEmails()
EMAIL_CONFIG_PATH = "conf/MOD_DailyReport_Email_Conf.json"
SUBJECT = f"MOD Daily Report {TODAY_STR}"
BODY = "Please find the attached daily report for Meno On-Demand.\n\n\nThis is an automated email, please do not reply."

# Excel Report Constants
CUSTOMER_DATA_HEADERS = ["Customer", "Total Active Orders", "Day 0", "Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6+"]
PRODUCT_DATA_HEADERS = ["Product ID", "Total Active Items",  "Day 0", "Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6+"]

# Notion Database IDs
JOB_DB_ID = "f11c954da24143acb6e2bf0254b64079"
ORDER_DB_ID = "d2747a287e974348870a636fbfa91e3e"

WORKSHEET_TITLE = f"MOD Daily Report - {TODAY_STR}"

# Alignment styles
C_ALIGN = Alignment(horizontal='center', vertical='center')
L_ALIGN = Alignment(horizontal='left', vertical='center')
R_ALIGN = Alignment(horizontal='right', vertical='center')

# Font styles
BOLD_FONT = Font(bold=True)

# Fill styles
LIGHT_GRAY_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

# Border styles
LTB_BORDER = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
RTB_BORDER = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
RT_BORDER = Border(right=Side(style='thin'), top=Side(style='thin'))
LT_BORDER = Border(left=Side(style='thin'), top=Side(style='thin'))
TB_BORDER = Border(top=Side(style='thin'), bottom=Side(style='thin'))
L_BORDER = Border(left=Side(style='thin'))
R_BORDER = Border(right=Side(style='thin'))
B_BORDER = Border(bottom=Side(style='thin'))
LB_BORDER = Border(left=Side(style='thin'), bottom=Side(style='thin'))
RB_BORDER = Border(right=Side(style='thin'), bottom=Side(style='thin'))

# Collects the largest width of each column based off of table header char count.
max_width = {}
DEFAULT_COLUMN_WIDTH = 15

# Data structure templates
CUSTOMER_DATA_TEMPLATE = {
    'total_active_orders': 0,
    0: 0,
    1: 0,
    2: 0,
    3: 0,
    4: 0,
    5: 0,
    6: 0,
}

PRODUCT_DATA_TEMPLATE = {
    'total_active_items': 0,
    0: 0,
    1: 0,
    2: 0,
    3: 0,
    4: 0,
    5: 0,
    6: 0,
}

status_tracker = {
    'Queued': 0,
    'Nest': 0,
    'Nesting': 0,
    'Print': 0,
    'Production': 0,
    'Packout': 0
}

# Orders currently in production.
ACTIVE_ORDERS_FILTER = {
    "and": [
        {
            "property": "Status",
            "select": {
                "does_not_equal": "Canceled"
            }
        },
        {
            "property": "Status",
            "select": {
                "does_not_equal": "Invoiced"
            }
        },
        {
            "property": "Status",
            "select": {
                "does_not_equal": "Shipped"
            }
        },
        {
            "property": "System status",
            "select": {
                "equals": "Active"
            }
        }
    ]
}

# Jobs currently in production.
ACTIVE_JOBS_FILTER = {
    "and": [
        {
            "property": "Job status",
            "select": 
            {
                "does_not_equal": "Canceled"
            }
        },
        {
            "property": "Job status",
            "select": {
                "does_not_equal": "Complete"
            }
        },
        {
            "property": "System status",
            "select": {
                "equals": "Active"
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Shipped"
                    }
                }
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Canceled"
                    }
                }
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Invoiced"
                    }
                }
            }
        }
    ]
}

ORDERS_SHIPPED_TODAY_FILTER = {
    "and": [
        {
            "property": "Shipped date",
            "date": {
                "equals": TODAY_STR
            }
        }
    ]
}

ORDERS_RECEIVED_TODAY = {
    "and": [
        {
            "timestamp": "created_time",
            "created_time": {
                "equals": TODAY_STR
            }
        },
        {
            "property": "System status",
            "select": {
                "equals": "Active"
            }
        }
    ]
}

ORDERS_SHIPPED_THIS_WEEK_FILTER = {
    "and": [
        {
            "property": "Shipped date",
            "date": {
                "this_week": {}
            }
        }
    ]
}

JOBS_IN_ERROR_FILTER = {
    "and": [
        {
            "property": "System status",
            "select": {
                "equals": "Error"
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Shipped"
                    }
                }
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Canceled"
                    }
                }
            }
        },
        {
            "property": "Order status",
            "rollup": {
                "any": {
                    "select": {
                        "does_not_equal": "Invoiced"
                    }
                }
            }
        }
    ]
}


def query_db(db_id, filter):
    logger.info(f"query_db() called for DB: {db_id}")
    response = notion_helper.query(db_id, content_filter=filter)
    
    if not response:
        response = []
    
    return response

def get_customer_data(data):
    """
    Extracts and processes customer data from the given Notion order data.
    Args:
        data (list): A list of dictionaries where each dictionary represents a page with 'properties', 'id', and 'created_time'.
    Returns:
        dict: A dictionary where the keys are customer names and the values are dictionaries containing customer data.
              The customer data includes:
              - 'total_active_orders': The total number of active orders for the customer.
              - A count of orders based on their age in days (0-6 days, with 6 representing 6 or more days).
    Notes:
        - The function calculates the age of each order in business days (Monday to Friday).
        - The 'created_time' is expected to be in the format "%Y-%m-%dT%H:%M:%S.%fZ".
        - The function uses a template dictionary CUSTOMER_DATA_TEMPLATE to initialize customer data.
    """
    
    logger.info("get_customer_data() called.")
    customer_data = {}
    
    # Iterate through each page of the order data
    for page in data:
        props = page['properties']
        pid = page['id']
        
        # Get customer name
        customer_name = notion_helper.return_property_value(props['Customer name'], pid)
        if not customer_name:
            continue
        
        # Convert created time to datetime object
        created_str = page['created_time'] #"2024-10-31T13:31:00.000Z"
        created_datetime = datetime.strptime(created_str, "%Y-%m-%dT%H:%M:%S.%fZ")
        
        # Calculate the age of the order in business days
        delta_days = 0
        current_day = created_datetime
        while current_day < TODAY:
            current_day += timedelta(days=1)
            if current_day.weekday() < 5:  # Monday to Friday are 0-4
                delta_days += 1
        
        # Initialize customer data if not present
        if customer_name not in customer_data:
            customer_data[customer_name] = CUSTOMER_DATA_TEMPLATE.copy()
            
        # Update customer data
        customer_data[customer_name]['total_active_orders'] += 1
        
        if delta_days < 7:
            customer_data[customer_name][delta_days] += 1
        else:
            customer_data[customer_name][6] += 1
        
    return customer_data

def get_product_data(data):
    """
    Extracts and processes product data from the given Notion job data.
    Args:
        data (list): A list of dictionaries where each dictionary represents a page with 'properties', 'id', and 'created_time'.
    Returns:
        dict: A dictionary where keys are product IDs and values are dictionaries containing product data, including:
            - 'total_active_items': Total quantity of active items.
            - Day-specific quantities for the past week (0-6 days).
    The function also updates a global status tracker with job statuses.
    Notes:
        - If 'Product ID' or 'Job status' is not found, the function continues to the next page.
        - If 'Quantity' is not found, it defaults to 1.
        - If 'Job status' is not found, it defaults to "Queued".
        - The function calculates the number of weekdays between the creation date and today.
    """
    
    logger.info("get_product_data() called.")
    product_data = {}
    
    # Iterate through each page of the job data
    for page in data:
        props = page['properties']
        pid = page['id']
        
        # Get product ID
        product_id = notion_helper.return_property_value(props['Product ID'], pid)
        if not product_id:
            continue
        
        # Default job status to "Queued"
        job_status = notion_helper.return_property_value(props['Job status'], pid)
        if not job_status:
            job_status = "Queued"
        
        # Get quantity, default to 1 if not found
        quantity = notion_helper.return_property_value(props['Quantity'], pid)
        if not quantity:
            quantity = 1
        
        # Convert created time to datetime object
        created_str = page['created_time'] #"2024-10-31T13:31:00.000Z"
        created_datetime = datetime.strptime(created_str, "%Y-%m-%dT%H:%M:%S.%fZ")
        
        # Calculate the age of the order in business days
        delta_days = 0
        current_day = created_datetime
        while current_day < TODAY:
            current_day += timedelta(days=1)
            if current_day.weekday() < 5:  # Monday to Friday are 0-4
                delta_days += 1

        # Initialize product data if not present
        if product_id not in product_data:
            product_data[product_id] = PRODUCT_DATA_TEMPLATE.copy()
            
        # Update product data
        product_data[product_id]['total_active_items'] += quantity
        
        if delta_days < 7:
            product_data[product_id][delta_days] += quantity
        else:
            product_data[product_id][6] += quantity
        
        status_tracker[job_status] += 1
        
    return product_data


def write_header(ws, orders_shipped_today, orders_shipped_this_week, job_in_error, orders_received_today):
    logger.info("write_header() called.")
    
    row_num = 1
    
    # Merge cells for each section
    ws.merge_cells(f"A{row_num}:B{row_num}")
    ws.merge_cells(f"D{row_num}:E{row_num}")
    ws.merge_cells(f"G{row_num}:H{row_num}")
    ws.merge_cells(f"J{row_num}:K{row_num}")
    
    # Assigned values to columns
    title_pairs = {
        "A": "Orders Received Today",
        "D": "Orders Shipped Today",
        "G": "Orders Shipped This Week",
        "J": "Jobs in Error"
    }
    
    value_pairs = {
        "C": orders_received_today,
        "F": orders_shipped_today,
        "I": orders_shipped_this_week,
        "L": job_in_error
    }
    
    # Write values to worksheet
    for col, title in title_pairs.items():
        ws[f"{col}{row_num}"] = title
        ws[f"{col}{row_num}"].font = BOLD_FONT
        ws[f"{col}{row_num}"].alignment = R_ALIGN
        ws[f'{col}{row_num}'].border = LTB_BORDER
        
        val_col = chr(ord(col) + 2)
        value = value_pairs[val_col]
        ws[f"{val_col}{row_num}"] = value
        ws[f"{val_col}{row_num}"].alignment = L_ALIGN
        ws[f"{val_col}{row_num}"].border = RTB_BORDER
        ws[f"{val_col}{row_num}"].font = BOLD_FONT
    
    logger.info("Finished write_header().")
    return ws, row_num


def write_table(ws, data, total, headers, row_num):
    """
    Writes a table to an Excel worksheet, starting at the specified row number.
    Alternates the background color of rows for better readability, and adjusts column widths based on the data.
    Set up to handle any number of columns and rows greater than 2.
    Iterates through columns by converting a character to ASCII value and adding an offset.
    Args:
        ws (Worksheet): The worksheet object where the table will be written.
        data (dict): A dictionary containing the data to be written. The keys are customer names and the values are dictionaries of column data.
        total (int or float): The total value to be written at the end of the table.
        headers (list): A list of column headers.
        row_num (int): The starting row number for writing the table.
    Returns:
        tuple: A tuple containing the worksheet object and the next row number after the table.
    """
    
    logger.info("write_table() called.")
    
    # Convert char to ASCII value. Allows for iteration through columns.
    col_value = ord("A")
    
    # No data to write
    if not data:
        return ws, row_num
    
    # Write headers, degine header_row for alternating background color
    header_row = row_num
    for i, header in enumerate(headers):
        
        col = chr(col_value + i)
        ws[f'{col}{row_num}'] = header
        ws[f'{col}{row_num}'].font = BOLD_FONT
        ws[f'{col}{row_num}'].alignment = C_ALIGN
        ws[f'{col}{row_num}'].fill = LIGHT_GRAY_FILL
        
        if i == 0:
            ws[f'{col}{row_num}'].border = LTB_BORDER
        elif i == len(CUSTOMER_DATA_HEADERS) - 1:
            ws[f'{col}{row_num}'].border = RTB_BORDER
        else:
            ws[f'{col}{row_num}'].border = TB_BORDER
            
        if col not in max_width:
            max_width[col] = DEFAULT_COLUMN_WIDTH
        elif len(str(header)) > max_width[col]:
            max_width[col] = len(str(header))
            
    # Write data
    background_offset = header_row % 2
    for row_index, (customer_t, values_t) in enumerate(data.items()):
        row_num += 1
        col = chr(col_value)
        customer = customer_t
        values = values_t
        
        # start column A
        ws[f'{col}{row_num}'] = customer
        ws[f'{col}{row_num}'].alignment = C_ALIGN
        if row_index == len(data) - 1:
            ws[f'{col}{row_num}'].border = LB_BORDER
        else:
            ws[f'{col}{row_num}'].border = L_BORDER
        
        # Alternates background color
        if row_num % 2 == background_offset:
            ws[f'{col}{row_num}'].fill = LIGHT_GRAY_FILL
        
        # start column B-*
        for i, (key_t, value_t) in enumerate(values.items()):
            col = chr(col_value + 1 + i)
            value = value_t
            
            print(f"Row: {row_num}, Col: {col}, Value: {value}")
            ws[f'{col}{row_num}'] = value
            ws[f'{col}{row_num}'].alignment = C_ALIGN
            
            # Alternates background color
            if row_num % 2 == background_offset:
                ws[f'{col}{row_num}'].fill = LIGHT_GRAY_FILL
            
            # Sets border for last column
            if i == len(CUSTOMER_DATA_HEADERS) - 2 and row_index == len(data) - 1:
                ws[f'{col}{row_num}'].border = RB_BORDER
            elif i == len(CUSTOMER_DATA_HEADERS) - 2:
                ws[f'{col}{row_num}'].border = R_BORDER
            elif row_index == len(data) - 1:
                ws[f'{col}{row_num}'].border = B_BORDER
    
    # Write total row
    row_num += 1
    ws[f'A{row_num}'] = "Total"
    ws[f'A{row_num}'].font = BOLD_FONT
    ws[f'A{row_num}'].alignment = R_ALIGN
    ws[f'A{row_num}'].border = LTB_BORDER
    
    ws[f'B{row_num}'] = total
    ws[f'B{row_num}'].font = BOLD_FONT
    ws[f'B{row_num}'].alignment = C_ALIGN
    ws[f'B{row_num}'].border = RTB_BORDER

    return ws, row_num

def write_tracker(ws, row_num):
    logger.info("write_tracker() called.")
    
    # Write headers
    ws[f"A{row_num}"] = "Job Status"
    ws[f"A{row_num}"].font = BOLD_FONT
    ws[f"A{row_num}"].alignment = C_ALIGN
    ws[f"A{row_num}"].fill = LIGHT_GRAY_FILL
    ws[f"A{row_num}"].border = LTB_BORDER
    
    ws[f"B{row_num}"] = "Count"
    ws[f"B{row_num}"].font = BOLD_FONT
    ws[f"B{row_num}"].alignment = C_ALIGN
    ws[f"B{row_num}"].fill = LIGHT_GRAY_FILL
    ws[f"B{row_num}"].border = RTB_BORDER
    
    # Iterate through each item in the status dict.
    for i, (status, count) in enumerate(status_tracker.items()):
        row_num += 1
        col = "A"
        ws[f"{col}{row_num}"] = status
        ws[f"{col}{row_num}"].alignment = C_ALIGN
        ws[f"{col}{row_num}"].border = L_BORDER
        
        if i == len(status_tracker) - 1:
            ws[f"{col}{row_num}"].border = LB_BORDER
        elif i == 0:
            ws[f"{col}{row_num}"].border = LT_BORDER
        else:
            ws[f"{col}{row_num}"].border = L_BORDER
        
        col = "B"
        ws[f"{col}{row_num}"] = count
        ws[f"{col}{row_num}"].alignment = C_ALIGN
        ws[f"{col}{row_num}"].border = R_BORDER
        
        if i == len(status_tracker) - 1:
            ws[f"{col}{row_num}"].border = RB_BORDER
        elif i == 0:
            ws[f"{col}{row_num}"].border = RT_BORDER
        else:
            ws[f"{col}{row_num}"].border = R_BORDER
    
    logger.info("Finished write_tracker().")
    return ws, row_num


def set_column_width(ws):
    logger.info("set_column_width() called.")
    for col, width in max_width.items():
        ws.column_dimensions[col].width = width + 2
    
    return ws


def build_xlsx(orders_shipped_today, orders_shipped_this_week, jobs_in_error, customer_data, product_data,
                total_orders, total_items, orders_received_today):
    logger.info("build_xlsx() called.")
    
    # Initialize workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_TITLE
    
    # Write total numbers in first row.
    ws, row_number = write_header(ws, orders_shipped_today, orders_shipped_this_week, jobs_in_error, orders_received_today)

    # Write customer order data
    ws, row_number = write_table(ws, customer_data, total_orders, CUSTOMER_DATA_HEADERS, row_number + 2)

    # Write active product data
    ws, row_number = write_table(ws, product_data, total_items, PRODUCT_DATA_HEADERS, row_number + 2)
    
    # Write job status information
    ws, row_number = write_tracker(ws, row_number + 2)
    
    # Set column width
    ws = set_column_width(ws)
    
    # Save workbook
    wb.save(SHEET_OUTPUT_PATH)
    

def main():
    logger.info("[START]")
    
    # Query data for reports
    active_orders = query_db(ORDER_DB_ID, ACTIVE_ORDERS_FILTER)
    active_jobs = query_db(JOB_DB_ID, ACTIVE_JOBS_FILTER)
    
    # Process data into dictionaries to easily build report
    customer_data = get_customer_data(active_orders)
    product_data = get_product_data(active_jobs)
    
    # Get quantities for report
    orders_shipped_today = len(query_db(ORDER_DB_ID, ORDERS_SHIPPED_TODAY_FILTER))
    orders_shipped_this_week = len(query_db(ORDER_DB_ID, ORDERS_SHIPPED_THIS_WEEK_FILTER))
    jobs_in_error = len(query_db(JOB_DB_ID, JOBS_IN_ERROR_FILTER))
    orders_received_today = len(query_db(ORDER_DB_ID, ORDERS_RECEIVED_TODAY))
    total_orders = len(active_orders)

    # Get total products from job/product quantity data
    total_items = 0
    for product, dict in product_data.items():
        total_items += dict['total_active_items']
        
    # Build Excel sheet
    build_xlsx(orders_shipped_today, orders_shipped_this_week, jobs_in_error,
              customer_data, product_data, total_orders, total_items, orders_received_today)
    
    # Send email
    automated_emails.send_email(EMAIL_CONFIG_PATH, SUBJECT, BODY, [SHEET_OUTPUT_PATH])
    
    logger.info("[END]")
    

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"Error in main(): {e}", exc_info=True)